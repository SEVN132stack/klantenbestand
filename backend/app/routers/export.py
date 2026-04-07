import io
from datetime import datetime
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..database import get_db
from ..models import Client
from ..auth import can_read

router = APIRouter()

BLAUW    = "FF185FA5"
LICHTBLAUW = "FFE6F1FB"
GROEN    = "FFE1F5EE"
GRIJS    = "FFF5F5F3"
WIT      = "FFFFFFFF"
ZWART    = "FF1A1A1A"

def header_style(cell, bg=BLAUW):
    cell.font = Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        bottom=Side(style="thin", color="FFFFFFFF"),
        right=Side(style="thin", color="FFFFFFFF"),
    )

def data_style(cell, bold=False, number_format=None, bg=WIT, align="left"):
    cell.font = Font(name="Arial", size=10, bold=bold, color=ZWART)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if number_format:
        cell.number_format = number_format

def euro(cell, bg=WIT, bold=False):
    data_style(cell, bold=bold, number_format='€#,##0.00;(€#,##0.00);"-"', bg=bg, align="right")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

async def get_clients(db: AsyncSession):
    result = await db.execute(select(Client).order_by(Client.naam))
    return result.scalars().all()

async def get_beschikking_totalen(db: AsyncSession):
    """Haal per client_id de totalen op uit de beschikkingen tabel."""
    result = await db.execute(text("""
        SELECT client_id,
               COALESCE(SUM(bedrag_beschikt), 0) as beschikt,
               COALESCE(SUM(gefactureerd), 0)    as gefact,
               COALESCE(SUM(betaald), 0)          as betaald,
               COUNT(*)                            as aantal
        FROM beschikkingen
        GROUP BY client_id
    """))
    totalen = {}
    for row in result.fetchall():
        totalen[str(row[0])] = {
            "beschikt": float(row[1]),
            "gefact":   float(row[2]),
            "betaald":  float(row[3]),
            "aantal":   int(row[4]),
        }
    return totalen

def fmt_date(d):
    if not d:
        return "—"
    try:
        return d.strftime("%d-%m-%Y")
    except Exception:
        return str(d)[:10]

def safe_float(v):
    try:
        return float(v) if v is not None else 0.0
    except Exception:
        return 0.0

# ─── Rapport 1: Cliëntenoverzicht ────────────────────────────────────────────
@router.get("/clienten")
async def export_clienten(db: AsyncSession = Depends(get_db), _=can_read):
    clients = await get_clients(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Clienten"
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 30

    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value = "Clientenoverzicht — Export {}".format(datetime.now().strftime("%d-%m-%Y"))
    title.font = Font(bold=True, size=13, name="Arial", color="FFFFFFFF")
    title.fill = PatternFill("solid", fgColor=BLAUW)
    title.alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Naam","BSN","Status","Klant","Locatie","Begeleider 1","Begeleider 2","Datum start","Einde beschikking","Datum sluiting","Opmerkingen"]
    for col, h in enumerate(headers, 1):
        header_style(ws.cell(row=2, column=col, value=h))

    for r, c in enumerate(clients, 3):
        bg = WIT if r % 2 == 1 else GRIJS
        row = [c.naam, c.bsn or "—", c.status, c.klant or "—", c.locatie or "—",
               c.begeleider_1 or "—", c.begeleider_2 or "—",
               fmt_date(c.datum_start), fmt_date(c.einde_beschikking), fmt_date(c.datum_sluiting),
               c.opmerkingen or ""]
        for col, val in enumerate(row, 1):
            data_style(ws.cell(row=r, column=col, value=val), bg=bg)

    set_col_widths(ws, [25,12,18,25,14,14,14,14,16,14,30])
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "clienten_{}.xlsx".format(datetime.now().strftime("%Y%m%d"))
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename={}".format(fname)})


# ─── Rapport 2: Financieel overzicht ─────────────────────────────────────────
@router.get("/financieel")
async def export_financieel(db: AsyncSession = Depends(get_db), _=can_read):
    clients = await get_clients(db)
    totalen = await get_beschikking_totalen(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Financieel"
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 30

    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "Financieel overzicht — Export {}".format(datetime.now().strftime("%d-%m-%Y"))
    title.font = Font(bold=True, size=13, name="Arial", color="FFFFFFFF")
    title.fill = PatternFill("solid", fgColor=BLAUW)
    title.alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Naam","Klant","Status","# Beschikkingen","Beschikt (€)","Gefactureerd (€)","Betaald (€)","Openstaand (€)"]
    for col, h in enumerate(headers, 1):
        header_style(ws.cell(row=2, column=col, value=h))

    for r, c in enumerate(clients, 3):
        bg = WIT if r % 2 == 1 else GRIJS
        t = totalen.get(str(c.id), {"beschikt":0,"gefact":0,"betaald":0,"aantal":0})
        openstaand = t["beschikt"] - t["gefact"]

        data_style(ws.cell(row=r, column=1, value=c.naam), bg=bg)
        data_style(ws.cell(row=r, column=2, value=c.klant or "—"), bg=bg)
        data_style(ws.cell(row=r, column=3, value=c.status), bg=bg)
        data_style(ws.cell(row=r, column=4, value=t["aantal"]), bg=bg, align="center")
        euro(ws.cell(row=r, column=5, value=t["beschikt"]), bg=bg)
        euro(ws.cell(row=r, column=6, value=t["gefact"]), bg=bg)
        euro(ws.cell(row=r, column=7, value=t["betaald"]), bg=bg)
        euro(ws.cell(row=r, column=8, value=openstaand), bg=bg)

    last = len(clients) + 2
    total_row = last + 1
    ws.row_dimensions[total_row].height = 18
    data_style(ws.cell(row=total_row, column=1, value="TOTAAL"), bold=True, bg=LICHTBLAUW)
    for col in [2, 3, 4]:
        data_style(ws.cell(row=total_row, column=col, value=""), bg=LICHTBLAUW)
    for col in [5, 6, 7, 8]:
        cell = ws.cell(row=total_row, column=col,
                       value="=SUM({}3:{}{})".format(get_column_letter(col), get_column_letter(col), last))
        euro(cell, bg=LICHTBLAUW, bold=True)

    set_col_widths(ws, [25,25,18,14,18,18,18,18])
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "financieel_{}.xlsx".format(datetime.now().strftime("%Y%m%d"))
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename={}".format(fname)})


# ─── Rapport 3: Zorgstatus ───────────────────────────────────────────────────
@router.get("/zorgstatus")
async def export_zorgstatus(db: AsyncSession = Depends(get_db), _=can_read):
    clients = await get_clients(db)
    totalen = await get_beschikking_totalen(db)
    wb = Workbook()

    for label, filter_status in [("In zorg", ["In zorg"]), ("Uit zorg", ["Uit Zorg"]), ("Overig", ["Aangemeld","In ZTO","Afronden","Nieuwe beschikking aanvragen"])]:
        ws = wb.create_sheet(title=label)
        subset = [c for c in clients if c.status in filter_status]

        ws.merge_cells("A1:G1")
        title = ws["A1"]
        title.value = "{} — {} clienten — {}".format(label, len(subset), datetime.now().strftime("%d-%m-%Y"))
        title.font = Font(bold=True, size=12, name="Arial", color="FFFFFFFF")
        title.fill = PatternFill("solid", fgColor=BLAUW)
        title.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 14
        ws.row_dimensions[2].height = 28

        headers = ["Naam","Klant","Begeleider","Datum start","Einde beschikking","Beschikt (€)","Gefactureerd (€)"]
        for col, h in enumerate(headers, 1):
            header_style(ws.cell(row=2, column=col, value=h))

        for r, c in enumerate(subset, 3):
            bg = WIT if r % 2 == 1 else GRIJS
            t = totalen.get(str(c.id), {"beschikt":0,"gefact":0,"betaald":0})
            data_style(ws.cell(row=r, column=1, value=c.naam), bg=bg)
            data_style(ws.cell(row=r, column=2, value=c.klant or "—"), bg=bg)
            data_style(ws.cell(row=r, column=3, value=c.begeleider_1 or "—"), bg=bg)
            data_style(ws.cell(row=r, column=4, value=fmt_date(c.datum_start)), bg=bg)
            data_style(ws.cell(row=r, column=5, value=fmt_date(c.einde_beschikking)), bg=bg)
            euro(ws.cell(row=r, column=6, value=t["beschikt"]), bg=bg)
            euro(ws.cell(row=r, column=7, value=t["gefact"]), bg=bg)

        set_col_widths(ws, [25,25,16,14,16,18,18])
        ws.freeze_panes = "A3"

    del wb["Sheet"]
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "zorgstatus_{}.xlsx".format(datetime.now().strftime("%Y%m%d"))
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename={}".format(fname)})


# ─── Rapport 4: Per klant ────────────────────────────────────────────────────
@router.get("/per-klant")
async def export_per_klant(db: AsyncSession = Depends(get_db), _=can_read):
    clients = await get_clients(db)
    totalen = await get_beschikking_totalen(db)
    klanten = sorted(set(c.klant for c in clients if c.klant))
    wb = Workbook()

    for klant in klanten:
        subset = [c for c in clients if c.klant == klant]
        ws = wb.create_sheet(title=klant[:31])

        ws.merge_cells("A1:G1")
        title = ws["A1"]
        title.value = "{} — {} clienten".format(klant, len(subset))
        title.font = Font(bold=True, size=12, name="Arial", color="FFFFFFFF")
        title.fill = PatternFill("solid", fgColor=BLAUW)
        title.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 14
        ws.row_dimensions[2].height = 28

        headers = ["Naam","Status","Begeleider","Datum start","Einde beschikking","Beschikt (€)","Gefactureerd (€)"]
        for col, h in enumerate(headers, 1):
            header_style(ws.cell(row=2, column=col, value=h))

        for r, c in enumerate(subset, 3):
            bg = WIT if r % 2 == 1 else GRIJS
            t = totalen.get(str(c.id), {"beschikt":0,"gefact":0,"betaald":0})
            data_style(ws.cell(row=r, column=1, value=c.naam), bg=bg)
            data_style(ws.cell(row=r, column=2, value=c.status), bg=bg)
            data_style(ws.cell(row=r, column=3, value=c.begeleider_1 or "—"), bg=bg)
            data_style(ws.cell(row=r, column=4, value=fmt_date(c.datum_start)), bg=bg)
            data_style(ws.cell(row=r, column=5, value=fmt_date(c.einde_beschikking)), bg=bg)
            euro(ws.cell(row=r, column=6, value=t["beschikt"]), bg=bg)
            euro(ws.cell(row=r, column=7, value=t["gefact"]), bg=bg)

        set_col_widths(ws, [25,18,16,14,16,18,18])
        ws.freeze_panes = "A3"

    if not klanten:
        wb.create_sheet("Leeg")
    del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "per_klant_{}.xlsx".format(datetime.now().strftime("%Y%m%d"))
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename={}".format(fname)})
