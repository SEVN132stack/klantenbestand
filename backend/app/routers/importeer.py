import io
from datetime import datetime
from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import JSONResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
import openpyxl
from ..database import get_db, AsyncSessionLocal
from ..models import User
from ..auth import get_current_user

router = APIRouter()

# Kolomkoppeling voor CLIENT velden
CLIENT_KOLOMMEN = {
    "BSN": "bsn",
    "Client naam": "naam",
    "Naam": "naam",
    "Geboortedatum": "geboortedatum",
    "Status Client": "status",
    "Status": "status",
    "Klant": "klant",
    "Organisatie": "klant",
    "Locatie": "locatie",
    "Begeleider 1": "begeleider_1",
    "Begeleider 2": "begeleider_2",
    "Datum start": "datum_start",
    "Eerste beschikking": "datum_start",
    "Datum sluiting": "datum_sluiting",
    "Einde sluiting": "datum_sluiting",
    "Uur/dagdeel per week": "uur_per_week",
    "Tijd": "uur_per_week",
    "Opmerkingen": "opmerkingen",
    "Laatst Gefactureerd": "laatste_gefactureerd",
    "Enquete gestuurd": "enquete_gestuurd",
    "Evaluatieformulier aanwezig?": "enquete_gestuurd",
}

# Kolomkoppeling voor BESCHIKKING velden
BESCHIKKING_KOLOMMEN = {
    "Einde beschikking": "datum_einde",
    "Huidige beschikking": "datum_einde",
    "Bedrag beschikt": "bedrag_beschikt",
    "Gefactureerd": "gefactureerd",
    "Betaald": "betaald",
}

CLIENT_DATUM = {"geboortedatum", "datum_start", "datum_sluiting"}
BESCHIKKING_DATUM = {"datum_einde", "datum_start_b"}
FLOAT_VELDEN = {"bedrag_beschikt", "gefactureerd", "betaald"}

def parse_datum(v):
    if v is None:
        return None
    if isinstance(v, float) and v != v:
        return None
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()[:10]
    if not s or s.lower() in ("nan", "none", "-", ""):
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def parse_float(v):
    if v is None:
        return None
    if isinstance(v, float) and v != v:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ".").strip())
    except Exception:
        return None

def parse_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s.lower() in ("nan", "none", "") else s

@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if user.role != "admin":
        return JSONResponse(status_code=403, content={"detail": "Alleen admins mogen importeren"})

    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return JSONResponse(status_code=400, content={"detail": "Alleen .xlsx of .xls bestanden zijn toegestaan"})

    try:
        inhoud = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(inhoud), data_only=True)
    except Exception as e:
        return JSONResponse(status_code=400, content={"detail": "Ongeldig bestand: " + str(e)})

    blad = None
    for naam in ["Client informatie", "Clienten", "Sheet1"]:
        if naam in wb.sheetnames:
            blad = wb[naam]
            break
    if blad is None:
        blad = wb.active

    rijen = list(blad.iter_rows(values_only=True))
    if not rijen:
        return JSONResponse(status_code=400, content={"detail": "Bestand is leeg"})

    # Detecteer headerrij
    header_idx = 0
    for i, rij in enumerate(rijen[:5]):
        if len([c for c in rij if c is not None and str(c).strip()]) >= 3:
            header_idx = i
            break

    headers = [str(c).strip() if c is not None else "" for c in rijen[header_idx]]

    # Bouw kolom-indices op voor client en beschikking
    client_map = {}
    beschikking_map = {}
    for i, h in enumerate(headers):
        if h in CLIENT_KOLOMMEN and CLIENT_KOLOMMEN[h] not in client_map:
            client_map[CLIENT_KOLOMMEN[h]] = i
        if h in BESCHIKKING_KOLOMMEN and BESCHIKKING_KOLOMMEN[h] not in beschikking_map:
            beschikking_map[BESCHIKKING_KOLOMMEN[h]] = i

    if "naam" not in client_map:
        return JSONResponse(status_code=400, content={
            "detail": "Kolom 'Client naam' niet gevonden. Beschikbare kolommen: " + ", ".join(h for h in headers if h)
        })

    # Haal beschikbare DB kolommen op
    try:
        res = await db.execute(text(
            "SELECT column_name FROM information_schema.columns WHERE table_name='clienten'"
        ))
        db_client_cols = {r[0] for r in res.fetchall()}
        res2 = await db.execute(text(
            "SELECT column_name FROM information_schema.columns WHERE table_name='beschikkingen'"
        ))
        db_besch_cols = {r[0] for r in res2.fetchall()}
    except Exception:
        db_client_cols = set(CLIENT_KOLOMMEN.values()) | {"naam"}
        db_besch_cols = set(BESCHIKKING_KOLOMMEN.values())

    toegevoegd = 0
    beschikkingen_toegevoegd = 0
    overgeslagen = 0
    fouten = []

    for rij_nr, rij in enumerate(rijen[header_idx + 1:], start=header_idx + 2):
        if all(c is None or str(c).strip() in ("", "nan") for c in rij):
            continue

        # Verwerk cliëntdata
        client_data = {}
        for veld, idx in client_map.items():
            if idx >= len(rij) or veld not in db_client_cols:
                continue
            w = rij[idx]
            if veld in CLIENT_DATUM:
                client_data[veld] = parse_datum(w)
            else:
                client_data[veld] = parse_str(w)

        naam = client_data.get("naam")
        if not naam:
            overgeslagen += 1
            continue

        # Verwerk beschikkingdata
        besch_data = {}
        for veld, idx in beschikking_map.items():
            if idx >= len(rij) or veld not in db_besch_cols:
                continue
            w = rij[idx]
            if veld in BESCHIKKING_DATUM:
                besch_data[veld] = parse_datum(w)
            elif veld in FLOAT_VELDEN:
                besch_data[veld] = parse_float(w)
            else:
                besch_data[veld] = parse_str(w)

        # Datum start van client ook als beschikking datum_start gebruiken
        if "datum_start" in client_data and client_data["datum_start"]:
            besch_data.setdefault("datum_start", client_data["datum_start"])

        heeft_beschikking = any(v is not None for v in besch_data.values())

        try:
            async with AsyncSessionLocal() as sess:
                async with sess.begin():
                    # Client aanmaken
                    velden = {k: v for k, v in client_data.items() if k in db_client_cols}
                    velden.setdefault("status", "Aangemeld")
                    kolommen = ", ".join(velden.keys())
                    params = ", ".join(":" + k for k in velden.keys())
                    await sess.execute(
                        text("INSERT INTO clienten ({}) VALUES ({})".format(kolommen, params)),
                        velden
                    )

                    # Client ID ophalen
                    result = await sess.execute(
                        text("SELECT id FROM clienten WHERE naam=:naam ORDER BY aangemaakt DESC LIMIT 1"),
                        {"naam": naam}
                    )
                    new_id = result.scalar()

                    if new_id:
                        # Auditlog
                        await sess.execute(
                            text("""INSERT INTO audit_log (client_id, client_naam, user_id, user_naam, type, actie)
                                    VALUES (:cid, :cnaam, :uid, :unaam, 'add', 'Client geimporteerd via Excel')"""),
                            {"cid": new_id, "cnaam": naam, "uid": str(user.id), "unaam": user.naam}
                        )

                        # Beschikking aanmaken als er data is
                        if heeft_beschikking:
                            b_velden = {k: v for k, v in besch_data.items() if k in db_besch_cols and v is not None}
                            b_velden["client_id"] = new_id
                            b_velden["volgnummer"] = 1
                            b_kolommen = ", ".join(b_velden.keys())
                            b_params = ", ".join(":" + k for k in b_velden.keys())
                            await sess.execute(
                                text("INSERT INTO beschikkingen ({}) VALUES ({})".format(b_kolommen, b_params)),
                                b_velden
                            )
                            beschikkingen_toegevoegd += 1

            toegevoegd += 1
        except Exception as e:
            fouten.append("Rij {}: {}".format(rij_nr, str(e)[:120]))
            overgeslagen += 1

    return JSONResponse(content={
        "toegevoegd": toegevoegd,
        "beschikkingen_toegevoegd": beschikkingen_toegevoegd,
        "overgeslagen": overgeslagen,
        "fouten": fouten[:10],
        "bericht": "{} clienten geimporteerd ({} met beschikking), {} overgeslagen".format(
            toegevoegd, beschikkingen_toegevoegd, overgeslagen
        ),
    })
